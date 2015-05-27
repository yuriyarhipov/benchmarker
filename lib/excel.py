from openpyxl import load_workbook
from lxml import etree
import xlrd

class Excel(object):
    filename = None

    def __init__(self, filename):
        self.filename = filename

    def get_data(self):
        data =[]
        if '.xlsx' in self.filename:
            wb = load_workbook(filename=self.filename, use_iterators=True)
            ws = wb.active

            for excel_row in ws.iter_rows():
                row = []
                for cell in excel_row:
                    if cell.value:
                        row.append(cell.value)
                    else:
                        row.append('')
                data.append(row)
        else:
            try:
                workbook = xlrd.open_workbook(self.filename)
                worksheet = workbook.sheet_by_index(0)
                data = []
                for curr_row in range(worksheet.nrows):
                    data_row = []
                    row = worksheet.row(curr_row)
                    for curr_cell in range(worksheet.ncols):
                        data_row.append(str(worksheet.cell_value(curr_row, curr_cell)))
                    data.append(data_row)
            except:
                tree = etree.parse(self.filename)
                root = tree.getroot()
                for worksheet in root:
                    if 'Worksheet' in worksheet.tag:
                        for table in worksheet:
                            if 'Table' in table.tag:
                                for row in table:
                                    data_row = []
                                    if 'Row' in row.tag:
                                        for cell in row:
                                            for d in cell:
                                                data_row.append(d.text)
                                    data.append(data_row)

        return data

    def get_competitors_template(self):
        return self.get_data()[1:]

    def get_data_set(self):
        return self.get_data()








